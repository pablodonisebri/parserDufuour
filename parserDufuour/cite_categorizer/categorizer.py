import ssl
from urllib.request import urlopen
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from fastapi import HTTPException,  Path
import html
import unicodedata
from urllib.request import urlopen
import reportlab

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


class CiteCategorizer:
    def __init__(self):
        self.cites = []
        self.a1_lect_libros = {"Gen", "Gn", "Ex", "Lev", "Num", "Dt", "Jos", "Jue", "Rut", "1Sa",
                               "2Sa", "1Re", "2Re", "1Par", "2Par", "Esd", "Neh", "Tob", "Jdt", "Est", "1Mac", "2Mac"}
        self.a2_lect_libros = {"Job", "Sal", "Prov", "Ecl", "Cant", "Sab", "Eclo", "Is", "Jer", "Lam",
                               "Bar", "Ez", "Dan", "Os", "Jl", "Am", "Abd", "Jon", "Miq", "Naj", "Hab", "Sof", "Ag", "Zac", "Mal"}
        self.a3_lect_libros = {"Rom", "1Cor", "2Cor", "Ga", "Ef", "Flp", "Col", "1Tes", "2Tes",
                               "1Tim", "2Tim", "Tit", "Flm", "Heb", "Sant", "1Pe", "2Pe", "1Jn", "2Jn", "3Jn", "Jds", "Ap"}
        self.ev_lect_libros = {"Mt", "Mc", "Lc", "Jn"}
        self.a1_lect, self.a2_lect, self.a3_lect, self.ev_lect = [], [], [], []

        # If the path to the directory is not set, it will create a new directory
        self.create_directories()

    def create_directories(self):
        import os
        if not os.path.exists("/var/tmp/lecturas"):
            os.makedirs("/var/tmp/lecturas")

    def normalize_word(self, word: str) -> str:
        # Need to identify the accion%20de%20gracias as accion de gracias
        word = word.replace("%20", " ")
        return word

    def get_list_lecturas(self):
        try:
            url = "https://hjg.com.ar/vocbib/"

            ssl_context = ssl._create_unverified_context()
            page = urlopen(url, context=ssl_context)

            html_bytes = page.read()
            # Decodificar como ISO-8859-1 porque así está la página
            html_str = html_bytes.decode("iso-8859-1", errors="replace")

            # Extraer el texto de los enlaces
            lista_palabras = re.findall(
                r"<a[^>]*>(.*?)</a>", html_str, re.IGNORECASE)

            # Decodificar entidades HTML y limpiar espacios
            lista_limpia = [html.unescape(p).replace(
                "\xa0", " ").strip() for p in lista_palabras]

            return lista_limpia

        except Exception as e:
            raise HTTPException(
                status_code=404,
                detail=f"Failed to fetch webpage for lista de palabras: {str(e)}. Please retry"
            )

    def fetch_webpage(self, palabra: str):
        try:
            url = f"https://hjg.com.ar/vocbib/art/{palabra}.html"

            ssl_context = ssl._create_unverified_context()
            page = urlopen(url, context=ssl_context)

            html_bytes = page.read()
            return html_bytes.decode("utf-8", errors="ignore")
        except Exception as e:
            raise HTTPException(
                status_code=404,
                detail=f"Failed to fetch webpage for '{palabra}': {str(e)}. Please check the word spelling or try again later."
            )

    def extract_cites(self, html):
        return re.findall(r"<cite>(.*?)</cite>", html)

    def normalize_cites(self, cites):
        normalized = []
        last_book = None
        for cite in cites:
            match = re.match(r"^(\d*[A-Za-z]+)", cite)
            if match:
                last_book = match.group(0)
            else:
                cite = last_book + cite  # Prepend last seen book
            normalized.append(cite)
        return normalized

    def classify_cite(self, cite):
        book = re.match(r"^\d*[A-Za-z]+", cite).group()
        if book in self.a1_lect_libros:
            self.a1_lect.append(cite)
        elif book in self.a2_lect_libros:
            self.a2_lect.append(cite)
        elif book in self.a3_lect_libros:
            self.a3_lect.append(cite)
        elif book in self.ev_lect_libros:
            self.ev_lect.append(cite)

    def pad_list(self, lst, length):
        return lst + [""] * (length - len(lst))

    def create_excel(self, palabra: str):
        file_path = os.path.join(
            "/var/tmp/lecturas", f"lecturas_{palabra}.xlsx")

        max_length = max(len(self.a1_lect), len(self.a2_lect),
                         len(self.a3_lect), len(self.ev_lect))

        df = pd.DataFrame({
            "1a Lectura": self.pad_list(self.a1_lect, max_length),
            "2a Lectura": self.pad_list(self.a2_lect, max_length),
            "3a Lectura": self.pad_list(self.a3_lect, max_length),
            "Evangelio": self.pad_list(self.ev_lect, max_length)
        })

        df.to_excel(file_path, index=False, engine="openpyxl")

        # === Apply formatting using openpyxl ===
        wb = load_workbook(filename=file_path)
        ws = wb.active

        # Define header colors (1 per column)
        header_fills = [
            PatternFill(start_color="FFD966", end_color="FFD966",
                        fill_type="solid"),  # Gold
            PatternFill(start_color="A9D08E", end_color="A9D08E",
                        fill_type="solid"),  # Light Green
            PatternFill(start_color="9DC3E6", end_color="9DC3E6",
                        fill_type="solid"),  # Light Blue
            PatternFill(start_color="F4B084", end_color="F4B084",
                        fill_type="solid"),  # Light Orange
        ]

        for col_idx, cell in enumerate(ws[1], start=0):
            if col_idx < len(header_fills):
                cell.fill = header_fills[col_idx]

        # Zebra stripe (light grey on even rows)
        even_fill = PatternFill(start_color="F2F2F2",
                                end_color="F2F2F2", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = even_fill

        wb.save(file_path)
        print(f"Excel file '{file_path}' created successfully!")

    def create_df(self, palabra: str):

        max_length = max(len(self.a1_lect), len(self.a2_lect),
                         len(self.a3_lect), len(self.ev_lect))

        df = pd.DataFrame({
            "1a Lectura": self.pad_list(self.a1_lect, max_length),
            "2a Lectura": self.pad_list(self.a2_lect, max_length),
            "3a Lectura": self.pad_list(self.a3_lect, max_length),
            "Evangelio": self.pad_list(self.ev_lect, max_length)
        })

        return df

    def df_to_pdf_a4(self, df, palabra):
        """
        Generate a PDF from a pandas DataFrame with 4 equal columns
        that fit the A4 page width with 1 inch margins.

        Parameters:
            df: pandas.DataFrame with exactly 4 columns
            pdf_path: output PDF file path
        """
        pdf_path = os.path.join(
            "/var/tmp/lecturas", f"lecturas_{palabra}.pdf")
        excel_path = os.path.join(
            "/var/tmp/lecturas", f"lecturas_{palabra}.xlsx ")
        if df.shape[1] != 4:
            raise ValueError("DataFrame must have exactly 4 columns")

        # Convert DataFrame to list (including header)
        data = [list(df.columns)] + df.values.tolist()

        PAGE_WIDTH, PAGE_HEIGHT = A4
        inch = 72  # points per inch

        left_margin = right_margin = inch
        usable_width = PAGE_WIDTH - left_margin - right_margin

        col_width = usable_width / 4
        col_widths = [col_width] * 4

        doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                                leftMargin=left_margin, rightMargin=right_margin)

        table = Table(data, colWidths=col_widths)

        style = TableStyle([
            # Header color (gold)
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#FFD966")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ])

        # Zebra stripes on even rows (starting row 1, after header)
        for i in range(1, len(data)):
            if i % 2 == 0:
                style.add('BACKGROUND', (0, i), (-1, i),
                          colors.HexColor("#F2F2F2"))

        table.setStyle(style)

        elements = [table]
        doc.build(elements)

    def run(self, palabra: str):
        # every time we get call, we need to restart the cites buffer.
        self.cites = []
        self.a1_lect, self.a2_lect, self.a3_lect, self.ev_lect = [], [], [], []

        html = self.fetch_webpage(palabra)
        if not html:
            print("No HTML content fetched. Exiting.")
            return

        cites = self.extract_cites(html)
        normalized_cites = self.normalize_cites(cites)

        for cite in normalized_cites:
            self.classify_cite(cite)

        # self.create_excel(palabra=palabra)
        df = self.create_df(palabra)
        self.df_to_pdf_a4(df, palabra)
